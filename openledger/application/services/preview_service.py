from __future__ import annotations

import csv
import io
import threading
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pdfplumber import open as pdf_open
from pypdf import PdfReader

from openledger.infrastructure.storage.files.preview_cache import pdf_preview_cache_path
from openledger.state import resolve_under_root
from openledger.infrastructure.workflow.runtime import make_paths

_PDF_RENDER_LOCK = threading.Lock()


def preview_table(
    root: Path,
    run_id: str,
    rel_path: str,
    *,
    limit: int,
    offset: int,
) -> dict[str, Any]:
    paths = make_paths(root, run_id)
    file_path = resolve_under_root(paths.run_dir, rel_path)
    if not file_path.exists() or not file_path.is_file():
        raise FileNotFoundError("not found")

    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        rows: list[dict[str, str]] = []
        columns: list[str] = []
        has_more = False
        with file_path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            columns = [str(name) for name in (reader.fieldnames or [])]
            seen = 0
            for row in reader:
                if seen < offset:
                    seen += 1
                    continue
                if len(rows) >= limit:
                    has_more = True
                    break
                rows.append({key: str(value or "") for key, value in row.items()})
        return {
            "columns": columns,
            "rows": rows,
            "offset": offset,
            "limit": limit,
            "has_more": has_more,
            "next_offset": offset + limit if has_more else None,
            "prev_offset": max(offset - limit, 0) if offset > 0 else None,
        }

    if suffix in {".xlsx", ".xls"}:
        if suffix == ".xls":
            raise ValueError("preview supports xlsx only")
        return _preview_xlsx(file_path, limit=limit, offset=offset)

    raise ValueError("preview supports csv/xlsx only")


def _preview_xlsx(file_path: Path, *, limit: int, offset: int) -> dict[str, Any]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return {
                "columns": [],
                "rows": [],
                "offset": offset,
                "limit": limit,
                "has_more": False,
                "next_offset": None,
                "prev_offset": max(offset - limit, 0) if offset > 0 else None,
            }

        columns: list[str] = []
        for idx, cell in enumerate(header):
            text = "" if cell is None else str(cell).strip()
            columns.append(text if text else f"col_{idx + 1}")

        rows: list[dict[str, str]] = []
        seen = 0
        has_more = False
        for row in rows_iter:
            if seen < offset:
                seen += 1
                continue
            if len(rows) >= limit:
                has_more = True
                break
            parsed: dict[str, str] = {}
            for idx, col in enumerate(columns):
                val = row[idx] if idx < len(row) else None
                parsed[col] = "" if val is None else str(val)
            rows.append(parsed)
            seen += 1

        return {
            "columns": columns,
            "rows": rows,
            "offset": offset,
            "limit": limit,
            "has_more": has_more,
            "next_offset": offset + limit if has_more else None,
            "prev_offset": max(offset - limit, 0) if offset > 0 else None,
        }
    finally:
        workbook.close()


def get_pdf_meta(root: Path, run_id: str, rel_path: str) -> dict[str, Any]:
    paths = make_paths(root, run_id)
    file_path = resolve_under_root(paths.run_dir, rel_path)
    if not file_path.exists() or not file_path.is_file():
        raise FileNotFoundError("not found")
    if file_path.suffix.lower() != ".pdf":
        raise ValueError("preview supports pdf only")
    reader = PdfReader(str(file_path))
    return {"page_count": len(reader.pages)}


def render_pdf_page(
    root: Path,
    run_id: str,
    rel_path: str,
    *,
    page: int,
    dpi: int,
) -> bytes:
    paths = make_paths(root, run_id)
    file_path = resolve_under_root(paths.run_dir, rel_path)
    if not file_path.exists() or not file_path.is_file():
        raise FileNotFoundError("not found")
    if file_path.suffix.lower() != ".pdf":
        raise ValueError("preview supports pdf only")

    mtime = file_path.stat().st_mtime
    cache_path = pdf_preview_cache_path(
        paths.run_dir,
        rel_path,
        mtime=mtime,
        page=page,
        dpi=dpi,
    )

    with _PDF_RENDER_LOCK:
        if cache_path.exists():
            return cache_path.read_bytes()
        with pdf_open(file_path) as pdf:
            if page > len(pdf.pages):
                raise ValueError("page out of range")
            image = pdf.pages[page - 1].to_image(resolution=dpi).original
            buf = io.BytesIO()
            image.save(buf, format="PNG")
            data = buf.getvalue()
        cache_path.write_bytes(data)
        return data


def read_stage_log(root: Path, run_id: str, stage_id: str) -> str:
    paths = make_paths(root, run_id)
    log_path = paths.run_dir / "logs" / f"{stage_id}.log"
    if not log_path.exists():
        raise FileNotFoundError("log not found")
    return log_path.read_text(encoding="utf-8", errors="replace")
