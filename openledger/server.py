from __future__ import annotations

import csv
import hashlib
import io
import json
import mimetypes
import os
import shutil
import threading
import webbrowser
from pathlib import Path
from time import perf_counter
from typing import Literal, cast
from uuid import uuid4

import uvicorn
from fastapi import Body, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response
from openpyxl import load_workbook
from pdfplumber import open as pdf_open
from pydantic import BaseModel, ConfigDict, Field, JsonValue, RootModel, model_validator
from pypdf import PdfReader

from .capabilities import (
    get_capabilities_payload,
    get_pdf_parser_health,
    list_source_support_matrix,
)
from .config import (
    global_classifier_write_path,
    resolve_global_classifier_config,
)
from .files import safe_filename
from .logger import (
    current_request_id,
    get_logger,
    reset_request_id,
    set_request_id,
    setup_logging,
)
from .parsers.pdf import list_pdf_modes
from .profiles import (
    add_bill_from_run,
    check_profile_integrity,
    clear_run_binding,
    create_profile,
    get_run_binding,
    list_profiles,
    load_profile,
    remove_bills,
    reimport_bill,
    set_run_binding,
    update_profile,
)
from .profile_review import build_profile_review
from .settings import Settings, load_settings
from .state import resolve_under_root
from .workflow import (
    WorkflowRunner,
    create_run,
    get_state,
    list_artifacts,
    list_runs,
    make_paths,
    save_state,
)

_PDF_RENDER_LOCK = threading.Lock()


class RequestModel(BaseModel):
    model_config = ConfigDict(extra="forbid")


class ResponseModel(BaseModel):
    model_config = ConfigDict(extra="ignore")


class JsonObjectPayload(RootModel[dict[str, JsonValue]]):
    pass


class ErrorResponse(ResponseModel):
    error: str


class HealthResponse(ResponseModel):
    ok: bool


class PdfModeItem(ResponseModel):
    id: str
    name: str


class PdfModesResponse(ResponseModel):
    modes: list[PdfModeItem]


class ParserDetectSampleCheckModel(ResponseModel):
    index: int
    expected_kind: str
    detected_kind: str
    ok: bool


class PdfParserHealthItemModel(ResponseModel):
    mode_id: str
    mode_name: str
    status: Literal["ok", "warning", "error"]
    kinds: list[str]
    filename_hints: list[str]
    sample_checks: list[ParserDetectSampleCheckModel]
    warnings: list[str]
    errors: list[str]


class PdfParserHealthSummaryModel(ResponseModel):
    total: int
    ok: int
    warning: int
    error: int


class PdfParserHealthResponseModel(ResponseModel):
    summary: PdfParserHealthSummaryModel
    parsers: list[PdfParserHealthItemModel]


class SourceSupportItemModel(ResponseModel):
    id: str
    name: str
    channel: str
    file_types: list[str]
    filename_hints: list[str]
    stage: str
    parser_mode: str
    support_level: Literal["stable", "beta", "planned"]
    notes: str


class SourceSupportResponse(ResponseModel):
    sources: list[SourceSupportItemModel]


class CapabilitiesPayloadModel(ResponseModel):
    generated_at: str
    source_support_matrix: list[SourceSupportItemModel]
    pdf_parser_health: PdfParserHealthResponseModel


class RunMetaModel(ResponseModel):
    id: str
    name: str
    status: str
    created_at: str


class RunsResponse(ResponseModel):
    runs: list[str]
    runs_meta: list[RunMetaModel]


class RunBindingModel(ResponseModel):
    run_id: str
    profile_id: str
    created_at: str = ""
    updated_at: str = ""


class ProfileBindingResponse(ResponseModel):
    run_id: str
    binding: RunBindingModel | None


class RunInputItemModel(ResponseModel):
    name: str
    path: str
    size: int


class StageStateModel(ResponseModel):
    id: str
    name: str
    status: str
    started_at: str
    ended_at: str
    log_path: str
    error: str


class ProfileArchiveStateModel(ResponseModel):
    status: Literal["ok", "failed"]
    profile_id: str
    run_id: str
    error: str | None = None
    updated_at: str | None = None


class RunOptionsStateModel(ResponseModel):
    pdf_mode: str | None = None
    classify_mode: Literal["llm", "dry_run"]
    period_mode: Literal["billing", "calendar"] | None = None
    period_day: int | None = None
    period_year: int | None = None
    period_month: int | None = None


class RunStateResponse(ResponseModel):
    run_id: str
    name: str | None = None
    status: str
    created_at: str
    updated_at: str
    cancel_requested: bool
    current_stage: str | None
    profile_archive: ProfileArchiveStateModel | None = None
    profile_binding: RunBindingModel | None = None
    inputs: list[RunInputItemModel]
    options: RunOptionsStateModel
    stages: list[StageStateModel]


class ArtifactItemModel(ResponseModel):
    path: str
    name: str
    exists: bool = True
    size: int | None = None


class ArtifactsResponse(ResponseModel):
    artifacts: list[ArtifactItemModel]


class FileItemModel(ResponseModel):
    path: str
    name: str
    exists: bool
    size: int | None = None


class StageIOResponse(ResponseModel):
    stage_id: str
    inputs: list[FileItemModel]
    outputs: list[FileItemModel]


class MatchReasonModel(ResponseModel):
    reason: str
    count: int


class MatchStatsResponse(ResponseModel):
    stage_id: str
    matched: int
    unmatched: int
    total: int
    match_rate: float
    unmatched_reasons: list[MatchReasonModel]


class CsvPreviewResponse(ResponseModel):
    columns: list[str]
    rows: list[dict[str, str]]
    offset: int
    limit: int
    has_more: bool
    next_offset: int | None
    prev_offset: int | None


class PdfPreviewMetaResponse(ResponseModel):
    page_count: int


class LogResponse(ResponseModel):
    text: str


class ProfilesResponse(ResponseModel):
    profiles: list["ProfileListItemModel"]


class ProfileListItemModel(ResponseModel):
    id: str
    name: str
    created_at: str
    updated_at: str
    bill_count: int


class ProfileBillTotalsModel(ResponseModel):
    sum_amount: float
    sum_expense: float
    sum_income: float
    sum_refund: float
    sum_transfer: float
    count: float
    net: float | None = None


class ProfileBillModel(ResponseModel):
    run_id: str
    period_key: str
    year: int | None
    month: int | None
    period_mode: str
    period_day: int
    period_start: str | None = ""
    period_end: str | None = ""
    period_label: str | None = ""
    cross_month: bool | None = False
    created_at: str
    updated_at: str
    outputs: dict[str, str] = Field(default_factory=dict)
    totals: ProfileBillTotalsModel | None = None
    category_summary: list[dict[str, JsonValue]] = Field(default_factory=list)


class ProfileModel(ResponseModel):
    id: str
    name: str
    created_at: str
    updated_at: str
    bills: list[ProfileBillModel]


class ProfileIntegrityIssueModel(ResponseModel):
    run_id: str
    period_key: str
    issue: str
    path: str | None = None


class ProfileIntegrityResultModel(ResponseModel):
    profile_id: str
    ok: bool
    issues: list[ProfileIntegrityIssueModel]


class ProfileReviewScopeModel(ResponseModel):
    profile_id: str
    profile_name: str
    data_source: Literal["profile_bills"]
    year: int | None = None
    months: int
    total_bills: int
    scoped_bills: int
    complete_period_bills: int
    unassigned_bills: int


class ProfileReviewOverviewModel(ResponseModel):
    total_expense: float
    total_income: float
    net: float
    period_count: int
    anomaly_count: int
    salary_income: float
    subsidy_income: float
    other_income: float


class ReviewMonthlyPointModel(ResponseModel):
    period_key: str
    year: int
    month: int
    expense: float
    income: float
    net: float
    tx_count: int
    mom_expense_rate: float | None = None
    yoy_expense_rate: float | None = None
    salary_income: float
    subsidy_income: float
    other_income: float


class ReviewYearlyPointModel(ResponseModel):
    year: int
    expense: float
    income: float
    net: float
    tx_count: int


class ReviewCategorySliceModel(ResponseModel):
    category_id: str
    category_name: str
    expense: float
    income: float
    tx_count: int
    share_expense: float


class ReviewAnomalyModel(ResponseModel):
    code: str
    severity: Literal["low", "medium", "high"]
    title: str
    period_key: str
    run_id: str
    message: str
    value: float | None = None
    baseline: float | None = None
    delta_rate: float | None = None


class ProfileReviewResponseModel(ResponseModel):
    scope: ProfileReviewScopeModel
    overview: ProfileReviewOverviewModel
    monthly_points: list[ReviewMonthlyPointModel]
    yearly_points: list[ReviewYearlyPointModel]
    category_slices: list[ReviewCategorySliceModel]
    anomalies: list[ReviewAnomalyModel]
    integrity_issues: list[ProfileIntegrityIssueModel]


class CreateRunPayload(RequestModel):
    name: str = ""


class CreateProfilePayload(RequestModel):
    name: str = Field(min_length=1, max_length=80)


class AddBillPayload(RequestModel):
    run_id: str = Field(min_length=1)
    period_year: int | None = None
    period_month: int | None = None


class RemoveBillsPayload(RequestModel):
    period_key: str | None = None
    run_id: str | None = None

    @model_validator(mode="after")
    def validate_selector(self) -> "RemoveBillsPayload":
        if not (self.period_key or self.run_id):
            raise ValueError("missing period_key or run_id")
        return self


class ReimportBillPayload(RequestModel):
    period_key: str = Field(min_length=1)
    run_id: str = Field(min_length=1)


class RunOptionsPatch(RequestModel):
    pdf_mode: str | None = None
    classify_mode: Literal["llm", "dry_run"] | None = None
    period_mode: Literal["billing", "calendar"] | None = None
    period_day: int | None = None
    period_year: int | None = None
    period_month: int | None = None
    profile_id: str | None = None


class StartRunPayload(RequestModel):
    stages: list[str] | None = None
    options: RunOptionsPatch = Field(default_factory=RunOptionsPatch)


class ResetPayload(RequestModel):
    scope: Literal["classify"] = "classify"


class ReviewUpdateItem(RequestModel):
    txn_id: str = Field(min_length=1)
    final_category_id: str | None = None
    final_note: str | None = None
    final_ignored: bool | None = None
    final_ignore_reason: str | None = None


class ReviewUpdatesPayload(RequestModel):
    updates: list[ReviewUpdateItem]


class ConfigWriteResponse(ResponseModel):
    ok: bool


class RunStartResponse(ResponseModel):
    ok: bool
    run_id: str


class CancelResponse(ResponseModel):
    ok: bool


class ResetResponse(ResponseModel):
    ok: bool
    scope: str


class ReviewUpdateResponse(ResponseModel):
    ok: bool
    updated: int


class SetProfileBindingPayload(RequestModel):
    profile_id: str = Field(min_length=1)


class SetProfileBindingResponse(ResponseModel):
    ok: bool
    binding: RunBindingModel


class UpdateProfilePayload(RequestModel):
    name: str | None = Field(default=None, max_length=80)


class UploadSavedItem(ResponseModel):
    name: str
    path: str
    size: int


class UploadResponse(ResponseModel):
    saved: list[UploadSavedItem]


class RunCreateResponse(RunStateResponse):
    pass


ProfilesResponse.model_rebuild()


class WorkflowContext:
    def __init__(self, root: Path) -> None:
        self.root = root
        self.runner = WorkflowRunner(root)
        self.settings = load_settings()
        setup_logging(self.settings)
        self.logger = get_logger()


def _preview_cache_path(
    run_dir: Path, rel_path: str, mtime: float, page: int, dpi: int
) -> Path:
    key = f"{rel_path}|{mtime:.3f}|{page}|{dpi}".encode("utf-8")
    digest = hashlib.sha1(key).hexdigest()[:16]
    cache_dir = run_dir / "preview"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / f"pdf_{digest}_p{page}_d{dpi}.png"


def _read_json_object(path: Path) -> dict[str, JsonValue]:
    raw_data = json.loads(path.read_text(encoding="utf-8"))
    payload = JsonObjectPayload.model_validate(raw_data)
    return payload.root


def _state_with_binding(root: Path, run_id: str) -> RunStateResponse:
    paths = make_paths(root, run_id)
    state_payload = get_state(paths)
    options_payload = dict(state_payload.get("options", {}))
    options_payload.pop("profile_id", None)
    state_payload["options"] = options_payload
    try:
        state_payload["profile_binding"] = get_run_binding(root, run_id)
    except FileNotFoundError:
        state_payload["profile_binding"] = None
    return RunStateResponse.model_validate(state_payload)


def _file_item(run_dir: Path, file_path: Path) -> FileItemModel:
    rel = str(file_path.resolve().relative_to(run_dir.resolve()))
    if file_path.exists() and file_path.is_file():
        size = file_path.stat().st_size
    else:
        size = None
    return FileItemModel(
        path=rel, name=file_path.name, exists=file_path.exists(), size=size
    )


def _glob_items(run_dir: Path, base: Path, pattern: str) -> list[FileItemModel]:
    return [_file_item(run_dir, p) for p in sorted(base.glob(pattern)) if p.is_file()]


def _one_item(run_dir: Path, p: Path) -> list[FileItemModel]:
    return [_file_item(run_dir, p)]


def _stage_io(root: Path, run_id: str, stage_id: str) -> StageIOResponse:
    paths = make_paths(root, run_id)
    run_dir = paths.run_dir
    inputs_dir = paths.inputs_dir
    out_dir = paths.out_dir
    cfg_dir = paths.config_dir

    if stage_id == "extract_pdf":
        return StageIOResponse(
            stage_id=stage_id,
            inputs=_glob_items(run_dir, inputs_dir, "*.pdf"),
            outputs=_glob_items(run_dir, out_dir, "*.transactions.csv"),
        )

    if stage_id == "extract_exports":
        return StageIOResponse(
            stage_id=stage_id,
            inputs=_glob_items(run_dir, inputs_dir, "*.xlsx")
            + _glob_items(run_dir, inputs_dir, "*.csv"),
            outputs=_one_item(run_dir, out_dir / "wechat.normalized.csv")
            + _one_item(run_dir, out_dir / "alipay.normalized.csv"),
        )

    if stage_id == "match_credit_card":
        cc_candidates = _glob_items(run_dir, out_dir, "*信用卡*.transactions.csv")
        if not cc_candidates:
            cc_candidates = _glob_items(run_dir, out_dir, "*.transactions.csv")
        return StageIOResponse(
            stage_id=stage_id,
            inputs=cc_candidates[:1]
            + _one_item(run_dir, out_dir / "wechat.normalized.csv")
            + _one_item(run_dir, out_dir / "alipay.normalized.csv"),
            outputs=_one_item(run_dir, out_dir / "credit_card.enriched.csv")
            + _one_item(run_dir, out_dir / "credit_card.unmatched.csv")
            + _one_item(run_dir, out_dir / "credit_card.match.xlsx")
            + _one_item(run_dir, out_dir / "credit_card.match_debug.csv"),
        )

    if stage_id == "match_bank":
        bank_candidates = _glob_items(run_dir, out_dir, "*交易流水*.transactions.csv")
        if not bank_candidates:
            bank_candidates = _glob_items(run_dir, out_dir, "*.transactions.csv")
        return StageIOResponse(
            stage_id=stage_id,
            inputs=bank_candidates
            + _one_item(run_dir, out_dir / "wechat.normalized.csv")
            + _one_item(run_dir, out_dir / "alipay.normalized.csv"),
            outputs=_one_item(run_dir, out_dir / "bank.enriched.csv")
            + _one_item(run_dir, out_dir / "bank.unmatched.csv")
            + _one_item(run_dir, out_dir / "bank.match.xlsx")
            + _one_item(run_dir, out_dir / "bank.match_debug.csv"),
        )

    if stage_id == "build_unified":
        return StageIOResponse(
            stage_id=stage_id,
            inputs=_one_item(run_dir, out_dir / "credit_card.enriched.csv")
            + _one_item(run_dir, out_dir / "credit_card.unmatched.csv")
            + _one_item(run_dir, out_dir / "bank.enriched.csv")
            + _one_item(run_dir, out_dir / "bank.unmatched.csv")
            + _one_item(run_dir, out_dir / "wechat.normalized.csv")
            + _one_item(run_dir, out_dir / "alipay.normalized.csv"),
            outputs=_one_item(run_dir, out_dir / "unified.transactions.csv")
            + _one_item(run_dir, out_dir / "unified.transactions.xlsx")
            + _one_item(run_dir, out_dir / "unified.transactions.all.csv")
            + _one_item(run_dir, out_dir / "unified.transactions.all.xlsx"),
        )

    if stage_id == "classify":
        classify_outputs = (
            _glob_items(run_dir, out_dir / "classify", "**/*")
            if (out_dir / "classify").exists()
            else []
        )
        return StageIOResponse(
            stage_id=stage_id,
            inputs=_one_item(run_dir, out_dir / "unified.transactions.csv")
            + _one_item(run_dir, cfg_dir / "classifier.json"),
            outputs=classify_outputs,
        )

    if stage_id == "finalize":
        return StageIOResponse(
            stage_id=stage_id,
            inputs=_one_item(run_dir, out_dir / "classify" / "unified.with_id.csv")
            + _one_item(run_dir, out_dir / "classify" / "review.csv")
            + _one_item(run_dir, cfg_dir / "classifier.json"),
            outputs=_one_item(run_dir, out_dir / "unified.transactions.categorized.csv")
            + _one_item(run_dir, out_dir / "unified.transactions.categorized.xlsx")
            + _one_item(run_dir, out_dir / "category.summary.csv")
            + _one_item(run_dir, out_dir / "pending_review.csv"),
        )

    return StageIOResponse(stage_id=stage_id, inputs=[], outputs=[])


def _count_csv_rows(
    path: Path, status_key: str = "match_status"
) -> tuple[int, dict[str, int]]:
    if not path.exists() or not path.is_file():
        return 0, {}
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        has_status = status_key in fieldnames
        count = 0
        reasons: dict[str, int] = {}
        for row in reader:
            count += 1
            if not has_status:
                continue
            status_value = str(row.get(status_key, "") or "").strip() or "unknown"
            reasons[status_value] = reasons.get(status_value, 0) + 1
    return count, reasons


def _preview_xlsx(file_path: Path, limit: int, offset: int) -> CsvPreviewResponse:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return CsvPreviewResponse(
                columns=[],
                rows=[],
                offset=offset,
                limit=limit,
                has_more=False,
                next_offset=None,
                prev_offset=max(offset - limit, 0) if offset > 0 else None,
            )

        columns: list[str] = []
        for idx, cell in enumerate(header):
            text = "" if cell is None else str(cell).strip()
            columns.append(text if text else f"col_{idx + 1}")

        rows: list[dict[str, str]] = []
        seen = 0
        has_more = False
        for row in rows_iter:
            if seen < offset:
                seen += 1
                continue
            if len(rows) >= limit:
                has_more = True
                break
            parsed: dict[str, str] = {}
            for idx, col in enumerate(columns):
                val = row[idx] if idx < len(row) else None
                parsed[col] = "" if val is None else str(val)
            rows.append(parsed)
            seen += 1

        return CsvPreviewResponse(
            columns=columns,
            rows=rows,
            offset=offset,
            limit=limit,
            has_more=has_more,
            next_offset=offset + limit if has_more else None,
            prev_offset=max(offset - limit, 0) if offset > 0 else None,
        )
    finally:
        workbook.close()


def _save_upload_files(
    paths_run_inputs: Path, files: list[UploadFile]
) -> list[UploadSavedItem]:
    paths_run_inputs.mkdir(parents=True, exist_ok=True)

    def pick_unique_name(name: str) -> str:
        dst = paths_run_inputs / name
        if not dst.exists():
            return name
        stem = Path(name).stem
        suffix = Path(name).suffix
        for index in range(1, 1000):
            candidate = f"{stem}_{index}{suffix}"
            if not (paths_run_inputs / candidate).exists():
                return candidate
        return f"{stem}_{os.getpid()}{suffix}"

    saved: list[UploadSavedItem] = []
    for upload in files:
        if not upload.filename:
            continue
        safe_name = safe_filename(upload.filename, default="upload.bin")
        target_name = pick_unique_name(safe_name)
        target_path = paths_run_inputs / target_name
        file_bytes = upload.file.read()
        target_path.write_bytes(file_bytes)
        saved.append(
            UploadSavedItem(
                name=target_name,
                path=f"inputs/{target_name}",
                size=target_path.stat().st_size,
            )
        )
    return saved


def create_app(root: Path) -> FastAPI:
    ctx = WorkflowContext(root)

    app = FastAPI(title="OpenLedger API", version="1.0.0")
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=False,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    def _extract_log_scope(path: str) -> tuple[str, str]:
        parts = [seg for seg in path.split("/") if seg]
        run_id = "-"
        stage_id = "-"
        if len(parts) >= 3 and parts[0] == "api" and parts[1] == "runs":
            run_id = parts[2] or "-"
        if (
            len(parts) >= 5
            and parts[0] == "api"
            and parts[1] == "runs"
            and parts[3] == "logs"
        ):
            stage_id = parts[4] or "-"
        if (
            len(parts) >= 6
            and parts[0] == "api"
            and parts[1] == "runs"
            and parts[3] == "stages"
        ):
            stage_id = parts[4] or "-"
        return run_id, stage_id

    def _request_logger(request: Request, request_id: str):
        run_id, stage_id = _extract_log_scope(request.url.path)
        return ctx.logger.bind(
            request_id=request_id,
            run_id=run_id,
            stage_id=stage_id,
            method=request.method,
            path=request.url.path,
        )

    @app.middleware("http")
    async def request_log_middleware(request: Request, call_next):
        request_id = (
            str(request.headers.get("x-request-id", "") or "").strip()
            or uuid4().hex[:16]
        )
        token = set_request_id(request_id)
        started = perf_counter()
        req_logger = _request_logger(request, request_id)
        try:
            response = await call_next(request)
        except Exception:
            duration_ms = (perf_counter() - started) * 1000
            req_logger.bind(status_code=500).opt(exception=True).error(
                f"request failed duration_ms={duration_ms:.2f}"
            )
            reset_request_id(token)
            raise

        duration_ms = (perf_counter() - started) * 1000
        response.headers["X-Request-Id"] = request_id
        status_code = int(response.status_code)
        message = (
            f"request completed status={status_code} duration_ms={duration_ms:.2f}"
        )
        if status_code >= 500:
            req_logger.bind(status_code=status_code).error(message)
        elif status_code >= 400:
            req_logger.bind(status_code=status_code).warning(message)
        else:
            req_logger.bind(status_code=status_code).info(message)
        reset_request_id(token)
        return response

    @app.exception_handler(HTTPException)
    async def http_exc_handler(request: Request, exc: HTTPException) -> JSONResponse:
        request_id = current_request_id()
        req_logger = _request_logger(request, request_id)
        detail = str(exc.detail)
        if exc.status_code >= 500:
            req_logger.bind(status_code=exc.status_code).error(
                f"http exception: {detail}"
            )
        elif exc.status_code >= 400:
            req_logger.bind(status_code=exc.status_code).warning(
                f"http exception: {detail}"
            )
        return JSONResponse(
            status_code=exc.status_code,
            content={"error": detail, "request_id": request_id},
        )

    @app.exception_handler(RequestValidationError)
    async def validation_exc_handler(
        request: Request, exc: RequestValidationError
    ) -> JSONResponse:
        request_id = current_request_id()
        req_logger = _request_logger(request, request_id)
        req_logger.bind(status_code=422).warning("request validation failed")
        return JSONResponse(
            status_code=422,
            content={
                "error": "request validation failed",
                "request_id": request_id,
                "details": exc.errors(),
            },
        )

    @app.exception_handler(Exception)
    async def generic_exc_handler(request: Request, exc: Exception) -> JSONResponse:
        request_id = current_request_id()
        req_logger = _request_logger(request, request_id)
        req_logger.bind(status_code=500).opt(exception=True).error(
            "unhandled exception"
        )
        return JSONResponse(
            status_code=500,
            content={"error": "internal server error", "request_id": request_id},
        )

    @app.get("/api/health", response_model=HealthResponse)
    async def health() -> HealthResponse:
        return HealthResponse(ok=True)

    @app.get("/api/parsers/pdf", response_model=PdfModesResponse)
    async def parsers_pdf() -> PdfModesResponse:
        return PdfModesResponse.model_validate({"modes": list_pdf_modes()})

    @app.get("/api/parsers/pdf/health", response_model=PdfParserHealthResponseModel)
    async def parsers_pdf_health() -> PdfParserHealthResponseModel:
        return PdfParserHealthResponseModel.model_validate(get_pdf_parser_health())

    @app.get("/api/sources/support", response_model=SourceSupportResponse)
    async def source_support() -> SourceSupportResponse:
        return SourceSupportResponse.model_validate(
            {"sources": list_source_support_matrix()}
        )

    @app.get("/api/capabilities", response_model=CapabilitiesPayloadModel)
    async def capabilities() -> CapabilitiesPayloadModel:
        return CapabilitiesPayloadModel.model_validate(get_capabilities_payload())

    @app.get("/api/config/classifier")
    async def get_global_classifier_config() -> dict[str, JsonValue]:
        cfg_path = resolve_global_classifier_config(root)
        if not cfg_path.exists():
            raise HTTPException(status_code=404, detail="config not found")
        return _read_json_object(cfg_path)

    @app.get("/api/runs", response_model=RunsResponse)
    async def get_runs() -> RunsResponse:
        run_ids = list_runs(root)
        meta_items: list[RunMetaModel] = []
        for run_id in run_ids:
            state = get_state(make_paths(root, run_id))
            meta_items.append(
                RunMetaModel(
                    id=run_id,
                    name=str(state.get("name", "") or ""),
                    status=str(state.get("status", "") or ""),
                    created_at=str(state.get("created_at", "") or ""),
                )
            )
        return RunsResponse(runs=run_ids, runs_meta=meta_items)

    @app.get("/api/runs/{run_id}", response_model=RunStateResponse)
    async def get_run(run_id: str) -> RunStateResponse:
        return _state_with_binding(root, run_id)

    @app.get(
        "/api/runs/{run_id}/profile-binding", response_model=ProfileBindingResponse
    )
    async def get_profile_binding(run_id: str) -> ProfileBindingResponse:
        try:
            binding = get_run_binding(root, run_id)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail="run not found") from exc
        parsed_binding = RunBindingModel.model_validate(binding) if binding else None
        return ProfileBindingResponse(run_id=run_id, binding=parsed_binding)

    @app.get("/api/profiles", response_model=ProfilesResponse)
    async def get_profiles() -> ProfilesResponse:
        return ProfilesResponse.model_validate({"profiles": list_profiles(root)})

    @app.get("/api/profiles/{profile_id}", response_model=ProfileModel)
    async def get_profile(profile_id: str) -> ProfileModel:
        try:
            return ProfileModel.model_validate(load_profile(root, profile_id))
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail="profile not found") from exc

    @app.get(
        "/api/profiles/{profile_id}/check", response_model=ProfileIntegrityResultModel
    )
    async def check_profile(profile_id: str) -> ProfileIntegrityResultModel:
        try:
            result = check_profile_integrity(root, profile_id)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail="profile not found") from exc
        return ProfileIntegrityResultModel.model_validate(result)

    @app.get(
        "/api/profiles/{profile_id}/review", response_model=ProfileReviewResponseModel
    )
    async def profile_review(
        profile_id: str,
        year: int | None = Query(default=None, ge=1900, le=2200),
        months: int = Query(default=12, ge=6, le=120),
    ) -> ProfileReviewResponseModel:
        try:
            result = build_profile_review(root, profile_id, year=year, months=months)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail="profile not found") from exc
        return ProfileReviewResponseModel.model_validate(result)

    @app.get("/api/runs/{run_id}/artifacts", response_model=ArtifactsResponse)
    async def get_artifacts(run_id: str) -> ArtifactsResponse:
        artifacts = list_artifacts(make_paths(root, run_id))
        return ArtifactsResponse.model_validate({"artifacts": artifacts})

    @app.get("/api/runs/{run_id}/artifact")
    async def download_artifact(
        run_id: str, path: str = Query(min_length=1)
    ) -> FileResponse:
        paths = make_paths(root, run_id)
        file_path = resolve_under_root(paths.run_dir, path)
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(status_code=404, detail="not found")
        mime_type = (
            mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        )
        return FileResponse(file_path, media_type=mime_type, filename=file_path.name)

    @app.get("/api/runs/{run_id}/preview", response_model=CsvPreviewResponse)
    async def preview_table(
        run_id: str,
        path: str = Query(min_length=1),
        limit: int = Query(default=50, ge=1, le=5000),
        offset: int = Query(default=0, ge=0),
    ) -> CsvPreviewResponse:
        paths = make_paths(root, run_id)
        file_path = resolve_under_root(paths.run_dir, path)
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(status_code=404, detail="not found")

        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            rows: list[dict[str, str]] = []
            columns: list[str] = []
            has_more = False
            with file_path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                columns = [str(name) for name in (reader.fieldnames or [])]
                seen = 0
                for row in reader:
                    if seen < offset:
                        seen += 1
                        continue
                    if len(rows) >= limit:
                        has_more = True
                        break
                    rows.append({key: str(value or "") for key, value in row.items()})
            return CsvPreviewResponse(
                columns=columns,
                rows=rows,
                offset=offset,
                limit=limit,
                has_more=has_more,
                next_offset=offset + limit if has_more else None,
                prev_offset=max(offset - limit, 0) if offset > 0 else None,
            )

        if suffix in {".xlsx", ".xls"}:
            if suffix == ".xls":
                raise HTTPException(
                    status_code=400, detail="preview supports xlsx only"
                )
            return _preview_xlsx(file_path, limit=limit, offset=offset)

        raise HTTPException(status_code=400, detail="preview supports csv/xlsx only")

    @app.get("/api/runs/{run_id}/logs/{stage_id}", response_model=LogResponse)
    async def stage_log(run_id: str, stage_id: str) -> LogResponse:
        paths = make_paths(root, run_id)
        log_path = paths.run_dir / "logs" / f"{stage_id}.log"
        if not log_path.exists():
            raise HTTPException(status_code=404, detail="log not found")
        return LogResponse(text=log_path.read_text(encoding="utf-8", errors="replace"))

    @app.get("/api/runs/{run_id}/config/classifier")
    async def get_run_classifier_config(run_id: str) -> dict[str, JsonValue]:
        cfg = make_paths(root, run_id).config_dir / "classifier.json"
        if not cfg.exists():
            raise HTTPException(status_code=404, detail="config not found")
        return _read_json_object(cfg)

    @app.get("/api/runs/{run_id}/stages/{stage_id}/io", response_model=StageIOResponse)
    async def stage_io(run_id: str, stage_id: str) -> StageIOResponse:
        return _stage_io(root=root, run_id=run_id, stage_id=stage_id)

    @app.get("/api/runs/{run_id}/stats/match", response_model=MatchStatsResponse)
    async def match_stats(
        run_id: str,
        stage: Literal["match_credit_card", "match_bank"] = Query(...),
    ) -> MatchStatsResponse:
        paths = make_paths(root, run_id)
        if stage == "match_credit_card":
            matched_path = paths.out_dir / "credit_card.enriched.csv"
            unmatched_path = paths.out_dir / "credit_card.unmatched.csv"
        else:
            matched_path = paths.out_dir / "bank.enriched.csv"
            unmatched_path = paths.out_dir / "bank.unmatched.csv"

        matched_count, _ = _count_csv_rows(matched_path)
        unmatched_count, unmatched_reasons = _count_csv_rows(unmatched_path)
        total = matched_count + unmatched_count
        match_rate = round(matched_count / total, 4) if total else 0.0
        reasons_sorted = sorted(
            unmatched_reasons.items(), key=lambda item: (-item[1], item[0])
        )
        reasons = [
            MatchReasonModel(reason=key, count=value) for key, value in reasons_sorted
        ]
        return MatchStatsResponse(
            stage_id=stage,
            matched=matched_count,
            unmatched=unmatched_count,
            total=total,
            match_rate=match_rate,
            unmatched_reasons=reasons,
        )

    @app.get(
        "/api/runs/{run_id}/preview/pdf/meta", response_model=PdfPreviewMetaResponse
    )
    async def pdf_meta(
        run_id: str, path: str = Query(min_length=1)
    ) -> PdfPreviewMetaResponse:
        paths = make_paths(root, run_id)
        file_path = resolve_under_root(paths.run_dir, path)
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(status_code=404, detail="not found")
        if file_path.suffix.lower() != ".pdf":
            raise HTTPException(status_code=400, detail="preview supports pdf only")
        reader = PdfReader(str(file_path))
        return PdfPreviewMetaResponse(page_count=len(reader.pages))

    @app.get("/api/runs/{run_id}/preview/pdf/page")
    async def pdf_page(
        run_id: str,
        path: str = Query(min_length=1),
        page: int = Query(default=1, ge=1),
        dpi: int = Query(default=120, ge=72, le=200),
    ) -> Response:
        paths = make_paths(root, run_id)
        file_path = resolve_under_root(paths.run_dir, path)
        if not file_path.exists() or not file_path.is_file():
            raise HTTPException(status_code=404, detail="not found")
        if file_path.suffix.lower() != ".pdf":
            raise HTTPException(status_code=400, detail="preview supports pdf only")

        mtime = file_path.stat().st_mtime
        cache_path = _preview_cache_path(paths.run_dir, path, mtime, page, dpi)

        with _PDF_RENDER_LOCK:
            if cache_path.exists():
                data = cache_path.read_bytes()
            else:
                with pdf_open(file_path) as pdf:
                    if page > len(pdf.pages):
                        raise HTTPException(status_code=400, detail="page out of range")
                    image = pdf.pages[page - 1].to_image(resolution=dpi).original
                    buf = io.BytesIO()
                    image.save(buf, format="PNG")
                    data = buf.getvalue()
                cache_path.write_bytes(data)

        return Response(content=data, media_type="image/png")

    @app.post("/api/runs", response_model=RunCreateResponse)
    async def create_run_api(
        payload: CreateRunPayload | None = Body(default=None),
    ) -> RunCreateResponse:
        paths = create_run(root)
        run_id = paths.run_dir.name
        run_name = payload.name.strip()[:80] if payload else ""
        if run_name:
            state = get_state(paths)
            state["name"] = run_name
            save_state(paths, state)
        ctx.logger.bind(run_id=run_id, stage_id="-").info("已创建 Run")
        return RunCreateResponse.model_validate(
            _state_with_binding(root, run_id).model_dump()
        )

    @app.post("/api/profiles", response_model=ProfileModel)
    async def create_profile_api(payload: CreateProfilePayload) -> ProfileModel:
        profile = create_profile(root, payload.name)
        return ProfileModel.model_validate(profile)

    @app.post("/api/profiles/{profile_id}/bills", response_model=ProfileModel)
    async def add_bill(profile_id: str, payload: AddBillPayload) -> ProfileModel:
        kwargs: dict[str, int | None] = {}
        if payload.period_year is not None or payload.period_month is not None:
            kwargs["period_year"] = payload.period_year
            kwargs["period_month"] = payload.period_month
        try:
            profile = add_bill_from_run(root, profile_id, payload.run_id, **kwargs)
        except FileNotFoundError as exc:
            raise HTTPException(
                status_code=404, detail="profile or run not found"
            ) from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return ProfileModel.model_validate(profile)

    @app.post("/api/profiles/{profile_id}/bills/remove", response_model=ProfileModel)
    async def remove_bill(profile_id: str, payload: RemoveBillsPayload) -> ProfileModel:
        try:
            profile = remove_bills(
                root, profile_id, period_key=payload.period_key, run_id=payload.run_id
            )
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail="profile not found") from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return ProfileModel.model_validate(profile)

    @app.post("/api/profiles/{profile_id}/bills/reimport", response_model=ProfileModel)
    async def reimport_bill_api(
        profile_id: str, payload: ReimportBillPayload
    ) -> ProfileModel:
        try:
            profile = reimport_bill(
                root, profile_id, period_key=payload.period_key, run_id=payload.run_id
            )
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return ProfileModel.model_validate(profile)

    @app.post("/api/runs/{run_id}/upload", response_model=UploadResponse)
    async def upload_files(
        run_id: str, request: Request, files: list[UploadFile] = File(...)
    ) -> UploadResponse:
        paths = make_paths(root, run_id)
        if not paths.run_dir.exists():
            raise HTTPException(status_code=404, detail="run not found")

        content_length = request.headers.get("content-length", "")
        if not content_length:
            raise HTTPException(status_code=400, detail="missing Content-Length")
        total_bytes = int(content_length)
        max_bytes = ctx.settings.max_upload_bytes
        if total_bytes > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"payload too large: {total_bytes} > {max_bytes}",
            )

        saved = _save_upload_files(paths.inputs_dir, files)
        state = get_state(paths)
        state["inputs"] = [item.model_dump() for item in saved]
        save_state(paths, state)
        ctx.logger.bind(run_id=run_id, stage_id="-").info(f"已上传文件数: {len(saved)}")
        return UploadResponse(saved=saved)

    @app.post("/api/runs/{run_id}/start", response_model=RunStartResponse)
    async def start_run(run_id: str, payload: StartRunPayload) -> RunStartResponse:
        options_payload = payload.options.model_dump(exclude_unset=True)
        legacy_profile = str(options_payload.get("profile_id") or "").strip()
        if "profile_id" in options_payload:
            options_payload.pop("profile_id")
            if legacy_profile:
                try:
                    set_run_binding(root, run_id, legacy_profile)
                except FileNotFoundError as exc:
                    raise HTTPException(
                        status_code=404, detail="run or profile not found"
                    ) from exc
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail=str(exc)) from exc
            else:
                try:
                    clear_run_binding(root, run_id)
                except FileNotFoundError:
                    pass
        ctx.runner.start(
            run_id,
            stages=payload.stages,
            options=cast(dict[str, object], options_payload),
        )
        ctx.logger.bind(run_id=run_id, stage_id="-").info(
            f"已请求启动 Run（stages={payload.stages or 'all'}）"
        )
        return RunStartResponse(ok=True, run_id=run_id)

    @app.post("/api/runs/{run_id}/cancel", response_model=CancelResponse)
    async def cancel_run(run_id: str) -> CancelResponse:
        ctx.runner.request_cancel(run_id)
        ctx.logger.bind(run_id=run_id, stage_id="-").warning("已通过 API 请求取消")
        return CancelResponse(ok=True)

    @app.post("/api/runs/{run_id}/reset", response_model=ResetResponse)
    async def reset_run(run_id: str, payload: ResetPayload) -> ResetResponse:
        if ctx.runner.is_running(run_id):
            raise HTTPException(status_code=409, detail="run is running")
        if payload.scope != "classify":
            raise HTTPException(
                status_code=400, detail=f"unknown scope: {payload.scope}"
            )

        paths = make_paths(root, run_id)
        state = get_state(paths)

        shutil.rmtree(paths.out_dir / "classify", ignore_errors=True)
        for artifact in [
            paths.out_dir / "unified.transactions.categorized.csv",
            paths.out_dir / "unified.transactions.categorized.xlsx",
            paths.out_dir / "category.summary.csv",
            paths.out_dir / "category.summary.xlsx",
            paths.out_dir / "pending_review.csv",
        ]:
            try:
                artifact.unlink()
            except FileNotFoundError:
                pass

        for stage in state.get("stages", []):
            stage_id = str(stage.get("id", ""))
            if stage_id in {"classify", "finalize"}:
                stage["status"] = "pending"
                stage["started_at"] = ""
                stage["ended_at"] = ""
                stage["error"] = ""

        state["status"] = "idle"
        state["current_stage"] = None
        state["cancel_requested"] = False
        save_state(paths, state)
        ctx.logger.bind(run_id=run_id, stage_id="-").info("重置完成（scope=classify）")
        return ResetResponse(ok=True, scope=payload.scope)

    @app.post("/api/runs/{run_id}/review/updates", response_model=ReviewUpdateResponse)
    async def update_review(
        run_id: str, payload: ReviewUpdatesPayload
    ) -> ReviewUpdateResponse:
        paths = make_paths(root, run_id)
        review_path = paths.out_dir / "classify" / "review.csv"
        if not review_path.exists():
            raise HTTPException(status_code=404, detail="review.csv not found")

        update_map: dict[str, ReviewUpdateItem] = {u.txn_id: u for u in payload.updates}
        temp_path = review_path.with_suffix(".csv.tmp")

        with review_path.open("r", encoding="utf-8", newline="") as f_in:
            reader = csv.DictReader(f_in)
            fieldnames = [str(name) for name in (reader.fieldnames or [])]
            if "txn_id" not in fieldnames:
                raise HTTPException(status_code=400, detail="review.csv missing txn_id")
            rows = list(reader)

        editable_fields = {
            "final_category_id",
            "final_note",
            "final_ignored",
            "final_ignore_reason",
        }
        editable_fields = {name for name in editable_fields if name in set(fieldnames)}

        for row in rows:
            txn_id = str(row.get("txn_id", "") or "").strip()
            if txn_id not in update_map:
                continue
            update_item = update_map[txn_id]
            value_map = {
                "final_category_id": update_item.final_category_id,
                "final_note": update_item.final_note,
                "final_ignored": update_item.final_ignored,
                "final_ignore_reason": update_item.final_ignore_reason,
            }
            for key in editable_fields:
                value = value_map.get(key)
                if value is None:
                    continue
                row[key] = str(value)

        with temp_path.open("w", encoding="utf-8", newline="") as f_out:
            writer = csv.DictWriter(f_out, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        temp_path.replace(review_path)

        ctx.logger.bind(run_id=run_id, stage_id="-").info(
            f"已更新 review.csv 行数={len(update_map)}"
        )
        return ReviewUpdateResponse(ok=True, updated=len(update_map))

    @app.put("/api/config/classifier", response_model=ConfigWriteResponse)
    async def update_global_classifier(
        payload: JsonObjectPayload,
    ) -> ConfigWriteResponse:
        cfg_path = global_classifier_write_path(root)
        cfg_path.parent.mkdir(parents=True, exist_ok=True)
        cfg_path.write_text(
            json.dumps(payload.root, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        ctx.logger.bind(run_id="-", stage_id="-").info("已更新全局分类器配置")
        return ConfigWriteResponse(ok=True)

    @app.put("/api/runs/{run_id}/config/classifier", response_model=ConfigWriteResponse)
    async def update_run_classifier(
        run_id: str, payload: JsonObjectPayload
    ) -> ConfigWriteResponse:
        cfg = make_paths(root, run_id).config_dir / "classifier.json"
        cfg.write_text(
            json.dumps(payload.root, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        ctx.logger.bind(run_id=run_id, stage_id="-").info("已更新本次 Run 的分类器配置")
        return ConfigWriteResponse(ok=True)

    @app.put("/api/runs/{run_id}/options", response_model=ConfigWriteResponse)
    async def update_options(
        run_id: str, payload: RunOptionsPatch
    ) -> ConfigWriteResponse:
        options_payload = payload.model_dump(exclude_unset=True)

        legacy_profile_value = options_payload.pop("profile_id", None)
        if legacy_profile_value is not None:
            profile_id = str(legacy_profile_value).strip()
            try:
                if profile_id:
                    set_run_binding(root, run_id, profile_id)
                else:
                    clear_run_binding(root, run_id)
            except FileNotFoundError as exc:
                raise HTTPException(
                    status_code=404, detail="run or profile not found"
                ) from exc
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=str(exc)) from exc

        paths = make_paths(root, run_id)
        state = get_state(paths)
        options = dict(state.get("options", {}))
        options.update(options_payload)
        options.pop("profile_id", None)
        state["options"] = options
        save_state(paths, state)
        ctx.logger.bind(run_id=run_id, stage_id="-").info(
            f"已更新 options: {options_payload}"
        )
        return ConfigWriteResponse(ok=True)

    @app.put(
        "/api/runs/{run_id}/profile-binding", response_model=SetProfileBindingResponse
    )
    async def update_profile_binding(
        run_id: str, payload: SetProfileBindingPayload
    ) -> SetProfileBindingResponse:
        try:
            binding = set_run_binding(root, run_id, payload.profile_id.strip())
        except FileNotFoundError as exc:
            raise HTTPException(
                status_code=404, detail="run or profile not found"
            ) from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return SetProfileBindingResponse(
            ok=True, binding=RunBindingModel.model_validate(binding)
        )

    @app.put("/api/profiles/{profile_id}", response_model=ProfileModel)
    async def update_profile_api(
        profile_id: str, payload: UpdateProfilePayload
    ) -> ProfileModel:
        try:
            profile = update_profile(
                root, profile_id, payload.model_dump(exclude_unset=True)
            )
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail="profile not found") from exc
        return ProfileModel.model_validate(profile)

    @app.get("/{full_path:path}", include_in_schema=False)
    async def fallback(full_path: str) -> Response:
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="not found")
        return HTMLResponse(
            "<html><body>"
            "<h3>OpenLedger server is running.</h3>"
            "<p>Frontend is not served by this backend. Run <code>pnpm install</code> and <code>pnpm dev</code> under <code>web/</code>.</p>"
            "</body></html>"
        )

    return app


def serve(
    root: Path, host: str = "127.0.0.1", port: int = 8000, open_browser: bool = True
) -> None:
    settings: Settings = load_settings()
    setup_logging(settings)
    app = create_app(root)
    url = f"http://{host}:{port}"
    logger = get_logger()
    logger.bind(run_id="-", stage_id="-").info(f"UI 服务地址 -> {url}")
    if open_browser:
        threading.Timer(0.5, lambda: webbrowser.open(url)).start()
    uvicorn.run(
        app,
        host=host,
        port=port,
        log_level=settings.log_level.lower(),
        log_config=None,
        access_log=False,
    )


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    settings = load_settings()
    serve(
        root=root,
        host=settings.host,
        port=settings.port,
        open_browser=settings.open_browser,
    )


if __name__ == "__main__":
    main()
